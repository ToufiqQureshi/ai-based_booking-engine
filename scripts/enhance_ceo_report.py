"""
Enhance the CEO report (docs/CEO_Report_Staybooker_Jun2026.xlsx).

Adds depth requested for stakeholders:
  - Executive Summary refreshed to point at the new sheets + architecture diagram.
  - Sheet 6 closes with a professional, question-proof note for stakeholders.
  - Sheet 7  — Engineering Plan by Role (day-to-day work: API optimisation, bug hunting).
  - Sheet 8  — Silent Bugs & Server Stability (the quiet bugs that crash a server).
  - Sheet 9  — Scaling to Large Data (DB -> Backend -> Frontend in milliseconds, never blank).

Style is read from the existing workbook so new sheets match exactly.
Run from repo root:  python scripts/enhance_ceo_report.py
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PATH = "docs/CEO_Report_Staybooker_Jun2026.xlsx"

# ---- House style (sampled from the existing workbook) -----------------------
NAVY = "0B3050"      # title bar
BLUE = "1F4E79"      # header row / section text
GREY = "555555"      # subtitle
RED = "FFC7CE"       # Critical
ORANGE = "FFE3C0"    # High
YELLOW = "FFF2CC"    # Medium
GREEN = "E2EFDA"     # Good / already protected

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def title_font():   return Font(bold=True, size=16, color="FFFFFF")
def sub_font():     return Font(size=10, color=GREY)
def section_font(): return Font(bold=True, size=12, color=BLUE)
def head_font():    return Font(bold=True, size=10, color="FFFFFF")
def body_font():    return Font(size=10)
def body_bold():    return Font(bold=True, size=10)

def fill(hexc):     return PatternFill("solid", fgColor=hexc)
TOPWRAP = Alignment(wrap_text=True, vertical="top")
CENTERWRAP = Alignment(wrap_text=True, vertical="center")

SEV_FILL = {"Critical": RED, "High": ORANGE, "Medium": YELLOW, "Low": GREEN, "Done": GREEN}


def title_block(ws, title, subtitle, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    c = ws.cell(1, 1, title); c.font = title_font(); c.fill = fill(NAVY)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26
    s = ws.cell(2, 1, subtitle); s.font = sub_font(); s.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 28


def section(ws, row, text, span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text); c.font = section_font()
    return row + 1


def header_row(ws, row, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row, i, h)
        c.font = head_font(); c.fill = fill(BLUE); c.alignment = CENTERWRAP; c.border = BORDER
    ws.row_dimensions[row].height = 30
    return row + 1


def body_row(ws, row, values, sev_col=None, bold_first=False):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row, i, v)
        c.font = body_bold() if (bold_first and i == 1) else body_font()
        c.alignment = TOPWRAP; c.border = BORDER
        if sev_col and i == sev_col and v in SEV_FILL:
            c.fill = fill(SEV_FILL[v]); c.alignment = CENTERWRAP
    return row + 1


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ============================================================================
wb = load_workbook(PATH)

# ---------------------------------------------------------------------------
# 1) REFRESH EXECUTIVE SUMMARY  (rebuild cleanly, keep the house style)
# ---------------------------------------------------------------------------
old_idx = wb.sheetnames.index("1. Executive Summary")
del wb["1. Executive Summary"]
ws = wb.create_sheet("1. Executive Summary", index=old_idx)
set_widths(ws, {"A": 32, "B": 120})
title_block(
    ws,
    "Staybooker Booking Engine — CEO Report (June 2026)",
    "What has been solved, what is in progress, what risks are coming, how the product stays fast and crash-proof at scale, and why we need a team — at a glance.",
    4,
)
ws.freeze_panes = "A5"
r = 4
header_row(ws, r, ["Topic", "Detail"]); r += 1

rows = [
    ("WHAT THIS REPORT CONTAINS", "", True),
    ("Sheet 2 — Problems Solved", "45 issues fixed so far (5 Critical, 24 High, 16 Medium). Each issue shows what would have happened to the business if it had reached production.", False),
    ("Sheet 3 — Features Built", "Major product features delivered alongside the bug fixing — multi-property chains, loyalty program, OTA rate intelligence, AI concierge, real-time updates, encrypted secrets storage.", False),
    ("Sheet 4 — Work In Progress", "Critical items currently being worked on.", False),
    ("Sheet 5 — Future Risks", "13 real-world scenarios that WILL come as the business grows (100 users at once, payment failures, DDoS attacks, data growing to GBs) — what protection already exists and what still needs to be built.", False),
    ("Sheet 6 — Team & Hiring", "What a Tech Lead / Backend / Frontend / Database developer actually does, in simple language — who owns which part of the product as it grows, with a suggested hiring order and a note for stakeholders.", False),
    ("Sheet 7 — Engineering Plan by Role (NEW)", "The detailed, day-to-day work each developer does as we grow — how APIs are kept fast and cheap (optimisation), and the constant bug-hunting that keeps the server stable. This is the 'what exactly will they do all day' answer.", False),
    ("Sheet 8 — Silent Bugs & Server Stability (NEW)", "The most dangerous bugs are the quiet ones — no error on screen, but the server slowly leaks, stalls or crashes. This sheet lists that class of risk, what already protects us, and why the whole software will NOT go blank.", False),
    ("Sheet 9 — Scaling to Large Data (NEW)", "When a hotel has 2–3 years of bookings (GBs of data), it must still travel database → backend → browser in milliseconds and the page must never freeze. This sheet shows the speed budget and the plan to hold it.", False),
    ("Architecture diagram (NEW)", "A full visual map of the system — Staybooker_Architecture.excalidraw at the project root — showing how a request flows from the guest's browser through Cloudflare, the backend, Redis and the database, and back. Useful when onboarding any new engineer or stakeholder.", False),
    ("", "", False),
    ("KEY NUMBERS", "", True),
    ("Issues fixed till date", "45 (source: work_log_tracker.csv — engineering tasks logged up to 16-Jun-2026)", False),
    ("Critical issues fixed", "5 — payment secret key leak, payment amount tampering, permanent admin token, AI assistant crash, staff privilege escalation.", False),
    ("Security audit coverage", "A full security/performance audit found 36 findings (1 Critical, 12 High, 14 Medium, 9 Low). The Critical issue and most High findings are fixed; the rest are tracked openly in Sheets 4 and 5.", False),
    ("Automated tests added", "277 test cases passing (security, payments, booking, role access, rate limits, Redis failover). These run automatically on every code change (CI/CD), plus dependency vulnerability (CVE) scanning.", False),
    ("Money saved (examples)", "Payment tampering fix: a guest could pay Rs.1 and get a Rs.10,000 booking confirmed. AI cost fix: a bug was sending traffic to an AI model 10x more expensive than needed on the public chatbot. Tax validation fix: invalid tax setup could produce wrong bills for every guest.", False),
    ("Stability posture", "The app is built to fail SAFE: if Redis drops it falls back to in-memory limits; if a data fetch fails the guest sees a Retry button, never a blank page; security and money checks 'fail closed' (deny), while the screen 'fails soft' (keeps working). Details in Sheets 8 and 9.", False),
    ("", "", False),
    ("ONE-LINE MESSAGE FOR THE CEO", "", True),
    ("Bottom line", "The product is stable today and hoteliers' data is safe. The hiring need comes from growth: as hotels, traffic and data scale up, each layer — database, backend, frontend — needs a dedicated owner so the speed, safety and crash-proofing stay exactly where they are today. This report now also spells out, in plain language, the ongoing engineering work that keeps it that way (Sheets 7–9).", False),
]
for topic, detail, bold in rows:
    if topic == "" and detail == "":
        r += 1; continue
    a = ws.cell(r, 1, topic); a.alignment = TOPWRAP; a.border = BORDER
    a.font = body_bold() if bold else body_font()
    if bold:
        a.fill = fill(GREEN)
    b = ws.cell(r, 2, detail); b.alignment = TOPWRAP; b.border = BORDER; b.font = body_font()
    if bold:
        b.fill = fill(GREEN)
    r += 1

# Move the rebuilt sheet to the front (create_sheet at index keeps order, but be safe)
wb.move_sheet("1. Executive Summary", offset=-(wb.sheetnames.index("1. Executive Summary")))

# ---------------------------------------------------------------------------
# 2) APPEND A STAKEHOLDER NOTE TO SHEET 6 (Team & Hiring)
# ---------------------------------------------------------------------------
ws6 = wb["6. Team & Hiring"]
r = ws6.max_row + 2
r = section(ws6, r, "A NOTE FOR STAKEHOLDERS (why this ask, framed plainly)", 4)
ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
note = ws6.cell(r, 1,
    "This is not a request to fix something broken — the product works and customer data is safe today. It is a request to PROTECT what has been built and to DE-RISK growth. "
    "Up to now one person has carried database, backend and frontend together. That is fine at the current size, but every new hotel multiplies the data, the traffic and the number of places a single quiet bug could cause a slow page or an outage. "
    "Each role below removes a specific, named risk to revenue and reputation. Hiring in the order shown means we always add the person who removes the biggest risk next — never headcount for its own sake.")
note.font = body_font(); note.alignment = TOPWRAP; note.border = BORDER
ws6.row_dimensions[r].height = 90
r += 2

r = section(ws6, r, "WHAT EACH HIRE RETURNS (in business terms)", 4)
hdr = ["Role", "The risk it removes", "What the business gets back", "If we keep delaying"]
header_row(ws6, r, hdr); r += 1
data = [
    ("Backend Dev", "Stuck payments, slow APIs under load, runaway costs, a crash on a peak day.",
     "Money never gets stuck (auto-refund + reconciliation); the booking path stays fast when 100 people book at once; AI and server bills stay predictable.",
     "A peak-day outage or a stuck-payment incident on the highest-revenue day — exactly when it hurts most."),
    ("Frontend Dev", "Blank/white screens, double bookings, slow pages that make guests leave.",
     "Every page recovers from a network blip with a Retry; no duplicate bookings; pages load in under ~2s, which directly lifts conversion.",
     "Lost bookings every day to slow or broken pages — a silent, daily revenue leak that is hard to even notice."),
    ("Database Dev", "Dashboards crawling once data reaches GBs; a backup that has never been test-restored.",
     "Queries stay in milliseconds at any data size (indexes, partitioning, archiving); backups are proven to actually restore.",
     "The product feels slower for every hotel as it succeeds — and an un-tested backup is a bet the business cannot afford to lose."),
    ("Tech Lead", "One person reviewing everything; no second pair of eyes on security or money code.",
     "A human review gate on every change — the single most effective catch for the silent bugs in Sheet 8 — plus a clear roadmap.",
     "Risk concentrates on one person; quality depends on no one ever having an off day."),
]
for role, risk, gain, delay in data:
    r = body_row(ws6, r, [role, risk, gain, delay], bold_first=True)
ws6.freeze_panes = "A4"

# ---------------------------------------------------------------------------
# 3) SHEET 7 — ENGINEERING PLAN BY ROLE
# ---------------------------------------------------------------------------
ws = wb.create_sheet("7. Engineering Plan by Role")
set_widths(ws, {"A": 26, "B": 50, "C": 46, "D": 40})
title_block(ws, "Engineering Plan by Role — What Each Developer Does, Day to Day",
            "The concrete, ongoing work behind 'why a team'. This answers, in plain language: how do we keep the APIs fast (optimisation), and what is the constant bug-hunting that keeps the server stable as we grow?", 4)
ws.freeze_panes = "A5"
r = 4

def role_table(ws, r, role_title, intro, rows):
    r = section(ws, r, role_title, 4)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(r, 1, intro); c.font = sub_font(); c.alignment = TOPWRAP; c.border = BORDER
    ws.row_dimensions[r].height = 32
    r += 1
    r = header_row(ws, r, ["Ongoing job", "What the work actually is", "Why it matters to the business", "What it prevents"])
    for row in rows:
        r = body_row(ws, r, row, bold_first=True)
    return r + 1

r = role_table(ws, r, "BACKEND DEVELOPER",
    "Owns the APIs and their link to the database — payments, booking logic, security, speed and cost. This is where 'API optimisation' and most server-crash prevention live.",
    [
        ("API optimisation", "Measure each endpoint's response time, cache hot reads in Redis, replace many small queries with one batched query (kill N+1), enforce pagination caps, and push slow work (emails, sync) into background tasks.", "Fast pages keep guests booking and keep the server bill low — speed is revenue.", "Timeouts, slow dashboards, and a database overloaded by repeated heavy queries."),
        ("Continuous bug-hunting", "Every new endpoint is read for the quiet failure modes: swallowed exceptions, wrong query cardinality, missing limits, fail-open checks. (See Sheet 8 for the catalogue.)", "These are the bugs that don't show an error but quietly crash or stall the server.", "Silent outages and 'the site is slow / down' incidents that nobody can explain."),
        ("Payment integrity", "Build and run the auto-refund flow and a daily reconciliation job that matches the Razorpay ledger against our bookings.", "Guest money is never stuck; disputes and chargebacks drop.", "Manual refunds, angry guests, and Razorpay disputes."),
        ("Security on every route", "Add the role gate and tenant (hotel_id) filter to every new endpoint, and verify amounts/ids server-side.", "Each new feature ships without opening a data-isolation or tampering hole.", "Cross-hotel data leaks and price/role tampering."),
        ("Cost control", "Cap LLM tokens, enforce per-hotel AI quotas, bound DB rows returned and background work.", "AI and infra costs stay predictable as usage grows.", "A surprise lakhs-level AI bill from a single abusive script."),
        ("Observability", "Structured logs (no PII), error-rate alerts, and dashboards so problems are seen before customers complain.", "We hear about issues from our alerts, not from an angry hotelier.", "Problems festering silently until they become outages."),
    ])

r = role_table(ws, r, "FRONTEND DEVELOPER",
    "Owns the React + TypeScript screens and their link to the backend APIs. This is real engineering — correct data, fast loads, no blank screens — not visual design.",
    [
        ("API integration hardening", "Every page gets a loading state, an error state and a Retry — and tolerates slow or partial data without breaking.", "A network blip never shows the guest a blank page; the booking continues.", "White screens mid-booking → closed tabs → lost bookings."),
        ("Page-speed optimisation", "Code-splitting, lazy loading, image optimisation, and trimming the JavaScript bundle; monitor Core Web Vitals.", "Every extra second of load time cuts conversion ~7% — speed is direct revenue.", "Guests bouncing off a slow page before they ever see a room."),
        ("Large-data rendering", "Virtual scrolling and paged lists so a hotel with thousands of bookings still renders instantly.", "The dashboard stays smooth even for the biggest, most valuable hotels.", "The browser freezing while it tries to draw thousands of rows."),
        ("Double-submit & concurrency guards", "Disable buttons on submit, add idempotent retries, warn on conflicting edits across tabs.", "No accidental duplicate bookings or double payments on slow networks.", "Two bookings / two charges from one impatient double-click."),
        ("Real-time UI", "Wire dashboards to the live WebSocket/SSE updates so new bookings appear without a refresh.", "The product feels modern and 'alive'; staff react instantly.", "Stale screens and missed bookings."),
    ])

r = role_table(ws, r, "DATABASE DEVELOPER",
    "Owns the data layer — speed, schema, backups and the guarantee that data flows out smoothly at ANY size. Becomes essential as data approaches GB scale; can start part-time.",
    [
        ("Indexing & query plans", "Add and tune indexes on the hottest columns (hotel_id, dates, status); read query plans and fix slow ones.", "The busiest queries stay in milliseconds as data grows.", "Searches and reports getting slower for everyone over time."),
        ("Partitioning & archiving", "Move old bookings/analytics to separate tables; partition large tables by date/hotel.", "Day-to-day queries only touch recent, small data — so they stay fast.", "Dashboards taking 15–20s to open once a hotel has years of data."),
        ("Migration discipline", "Introduce Alembic so schema and index changes deploy in a controlled, repeatable way.", "Production never silently misses an index or a schema change.", "Indexes present in dev but missing in production → mystery slowness."),
        ("Connection pooling", "Add a pooler (PgBouncer) and tune pool size × workers × replicas against the Supabase limit.", "50–100 hotels can onboard without the database running out of connections.", "All tenants slowing down at once when connections run out."),
        ("Backups & restore drills", "Run real restore drills, set up point-in-time recovery, write a disaster-recovery runbook.", "Backups are proven to actually work before we ever need them.", "Discovering a backup is unusable on the worst possible day."),
        ("Isolation on schema changes", "Review every new table for Row Level Security and per-hotel isolation.", "The database itself keeps refusing to mix one hotel's data with another's.", "A schema change quietly opening a cross-tenant leak."),
    ])

r = role_table(ws, r, "TECH LEAD",
    "Responsible for the whole software — decides what each developer builds and when, reviews everything before release, and owns delivery and security posture.",
    [
        ("Roadmap & prioritisation", "Decide what gets built in what order; sequence work so the biggest risk is always removed next.", "Effort always goes to the highest-value, highest-risk item first.", "Busywork while a real risk sits unaddressed."),
        ("Code review on every change", "Be the human gate that reads every change for the silent-bug patterns in Sheet 8.", "The single most effective catch for the bugs that don't show an error.", "A quiet bug reaching production with no second pair of eyes."),
        ("Release & incident response", "Own the CI/CD pipeline, the deploy process, and the plan for when something breaks.", "Deploys are predictable and incidents are handled calmly, not in a panic.", "Risky deploys and chaotic, slow incident recovery."),
        ("Security posture", "Schedule penetration tests, drive the audit follow-ups, plan a future bug-bounty.", "Security stays ahead of attackers as the attack surface grows.", "Security drifting as new endpoints are added faster than they're reviewed."),
    ])

# ---------------------------------------------------------------------------
# 4) SHEET 8 — SILENT BUGS & SERVER STABILITY
# ---------------------------------------------------------------------------
ws = wb.create_sheet("8. Silent Bugs & Stability")
set_widths(ws, {"A": 30, "B": 42, "C": 40, "D": 40, "E": 18})
title_block(ws, "Silent Bugs & Server Stability — The Quiet Risks That Crash a Server",
            "The dangerous bugs are the ones with NO error on screen — the server slowly leaks memory, stalls, or crashes while everything looks fine. Below is that class of risk, what already protects us, and the ongoing work that keeps it safe. Most of these we have already met and fixed once (the S-xx references point to Sheet 2).", 5)
ws.freeze_panes = "A5"
r = 4
r = header_row(ws, r, ["Silent risk (plain language)", "How it quietly hurts the server", "Already protected", "Continuous work needed", "Owner"])
silent = [
    ("Swallowed errors (a background job fails but no one is told)", "Confirmation emails and channel-sync silently skip — guests get nothing, OTAs double-book — with zero error shown. We hit exactly this (S-12).", "Background-task pattern fixed; engineering rule bans bare 'except: pass'; errors must be logged or re-raised.", "Read every new background job in code review; alert on task failures.", "Backend + Tech Lead"),
    ("Query returns more rows than expected (cardinality crash)", "An endpoint that assumes one row gets several and throws a 500 only for certain data — invisible until that exact data appears.", "Rule to use .first() not scalar_one_or_none() where multiple rows are possible; tests around it.", "Review every new query for this pattern; add tests for multi-row cases.", "Backend"),
    ("Unbounded query (no limit) loads everything into memory", "A big hotel's list endpoint tries to load 100k rows and the server runs out of memory and crashes (S-13, S-43).", "Pagination (limit + offset) is mandatory on list endpoints, with hard caps; indexes on hot columns.", "Enforce a limit on EVERY new list endpoint; reject unbounded date ranges.", "Backend + Database"),
    ("'Fail-open' safety check (limit turns OFF when Redis is down)", "During a Redis blip the AI-spend limit silently passes everything — runaway cost or overload until someone notices (S-39, S-44).", "Quotas now 'fail closed'; an in-process cap takes over; Redis self-heals after a 30s back-off.", "Audit every new limit/lock to confirm it fails closed, not open.", "Backend"),
    ("N+1 queries (one page quietly fires hundreds of DB calls)", "Looks fine on small data; as data grows the database is hammered and every page slows for everyone.", "Batch loading with selectinload / single GROUP BY queries.", "Review new endpoints for query count; watch slow-query logs.", "Backend + Database"),
    ("Connection-pool exhaustion under load", "At 100 simultaneous bookings the pool empties and every request stalls — the whole site appears down (F-01, F-09).", "Env-driven pool sizing; documented pool math (SCALE_TUNING.md); row-level locks on inventory.", "Load testing, a connection pooler (PgBouncer), auto-scaling rules.", "Backend + Database + DevOps"),
    ("Stale cache serves wrong data", "Cache isn't cleared after a booking/cancellation, so a sold room shows as 'available' and gets double-sold (S-21).", "Cache invalidation + a version counter bumped after every booking/cancellation.", "Invalidation discipline on every new cached path.", "Backend"),
    ("Uncaught frontend error blanks the whole page", "One failed fetch with no error boundary turns the entire screen white mid-booking — the guest just leaves (S-36).", "Error + Retry UI on the booking pages; add-ons load separately so their failure can't blank the page.", "Same hardening on every dashboard page; a top-level error boundary.", "Frontend"),
    ("Resource leak (threads / streams pile up)", "Long-lived connections or threads are never released; the server slowly bloats and dies hours later — no obvious cause.", "Streaming connections have lifetime caps; web search runs in a thread with an 8s hard timeout.", "Monitor memory/thread counts; cap every new long-lived resource.", "Backend"),
    ("Unvalidated input / bad settings saved", "Negative tax, inverted slabs or a divide-by-zero produce wrong bills for every guest — silently, until accounts notice (S-41).", "Server-side validation rejects bad values; calculators guarded against NaN.", "Validate every new settings/input form server-side, not just in the browser.", "Backend + Frontend"),
]
for row in silent:
    r = body_row(ws, r, row, bold_first=True)

r += 1
r = section(ws, r, "WHY THE WHOLE SOFTWARE WILL NOT GO BLANK (the safety net)", 5)
for line in [
    "Layered fallbacks: if Redis crashes, the app transparently falls back to an in-memory cache — it never crashes because of Redis.",
    "No blank screens: a failed data fetch shows a clear message with a Retry button; optional pieces (add-ons) load separately so one failure can't take down the page.",
    "Fail closed on money & security, fail soft on the screen: anything touching payment, roles or another hotel's data DENIES on doubt; anything touching the display keeps working in a reduced mode.",
    "Tested on every change: 277 automated tests (including Redis-failover and cross-tenant attack tests) run before any code can merge.",
    "The 'safer side' principle the CEO asked for: when in doubt we choose the cautious behaviour — reject the risky action, keep the guest's screen usable, and log it for a human to review.",
]:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(r, 1, "•  " + line); c.font = body_font(); c.alignment = TOPWRAP; c.border = BORDER
    c.fill = fill(GREEN)
    ws.row_dimensions[r].height = 30
    r += 1

# ---------------------------------------------------------------------------
# 5) SHEET 9 — SCALING TO LARGE DATA
# ---------------------------------------------------------------------------
ws = wb.create_sheet("9. Scaling to Large Data")
set_widths(ws, {"A": 24, "B": 40, "C": 16, "D": 44, "E": 40, "F": 18})
title_block(ws, "Scaling to Large Data — Database → Backend → Browser in Milliseconds",
            "When a hotel has 2–3 years of bookings (GBs of data), the page must still load in milliseconds and must never freeze or go blank. This is the speed 'budget' for the journey of one request, what already holds it, and what we build to keep holding it as data grows.", 6)
ws.freeze_panes = "A5"
r = 4
r = header_row(ws, r, ["Stage of the journey", "What happens here", "Speed target", "In place today", "To build for GB-scale", "Owner"])
journey = [
    ("1. Database", "Find and return the right rows (bookings, rates, guests) for this hotel only.", "< 50 ms / query", "Indexes on hot columns; Row Level Security; pagination; composite index on the busiest availability query.", "Table partitioning by date/hotel; archive old bookings; read replicas for reports.", "Database"),
    ("2. Backend API", "Shape the rows into a clean API response, applying security and business rules.", "< 100 ms", "Redis caching on heavy aggregations; mandatory pagination; lazy heavy imports; batched queries (no N+1).", "Pre-computed report tables; response compression; background pre-warming of common queries.", "Backend"),
    ("3. Network & cache", "Carry the response to the browser, reusing cached copies where safe.", "< 100 ms (cache hit ~0)", "React Query caches GETs for 5 min; public reads cached in Redis.", "Cloudflare edge caching; ETags/conditional requests so unchanged data isn't re-sent.", "Backend + Frontend"),
    ("4. Browser render", "Draw the data on screen — lists, charts, the booking calendar.", "< 1 s, no freeze", "Paged lists; loading + error + Retry states on the booking pages.", "Virtual scrolling for huge lists; code-splitting & lazy loading; skeleton loaders.", "Frontend"),
    ("TOTAL (guest experience)", "From click to a fully usable screen.", "Well under ~2 s", "Holds comfortably at today's data sizes.", "Stays under ~2 s at GB-scale once the above are in place.", "All three roles"),
]
for row in journey:
    r = body_row(ws, r, row, bold_first=True)

r += 1
r = section(ws, r, "AND IT STAYS ON THE SAFER SIDE AT SCALE", 6)
for line in [
    "Even when data is huge, the app never tries to load everything at once — the database sends rows in pages, the backend serves them in chunks, and the frontend draws only what's on screen.",
    "If any stage is slow or fails, the guest sees a Retry — never a frozen tab or a white page.",
    "Money and security checks never get 'skipped to go faster': they stay strict (fail closed) no matter the load. Only the display degrades gracefully, never the safety.",
    "Costs stay bounded as data grows: capped result sizes, caching, and per-hotel quotas mean more data does not mean a runaway bill.",
]:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(r, 1, "•  " + line); c.font = body_font(); c.alignment = TOPWRAP; c.border = BORDER
    c.fill = fill(GREEN)
    ws.row_dimensions[r].height = 30
    r += 1

# ---------------------------------------------------------------------------
wb.save(PATH)
print("Saved:", PATH)
print("Sheets:", wb.sheetnames)
