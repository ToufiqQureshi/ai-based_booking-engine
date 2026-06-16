"""
- Rebuild Sheet 7 (Engineering Plan by Role): correct framing (Frontend = TypeScript
  engineering that talks to Backend APIs; Backend APIs talk to the Database — NOT design),
  with richer day-to-day detail for every role.
- Add a realistic closing note at the end of Sheet 5 (Future Risks).
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PATH = "docs/CEO_Report_Staybooker_Jun2026.xlsx"
NAVY, BLUE, GREY, GREEN = "0B3050", "1F4E79", "555555", "E2EFDA"
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP = Alignment(wrap_text=True, vertical="top")
CEN = Alignment(wrap_text=True, vertical="center")
def F(sz=10, b=False, c=None): return Font(size=sz, bold=b, color=c)
def fill(c): return PatternFill("solid", fgColor=c)

wb = load_workbook(PATH)

# ===========================================================================
# Rebuild Sheet 7
# ===========================================================================
idx = wb.sheetnames.index("7. Engineering Plan by Role")
del wb["7. Engineering Plan by Role"]
ws = wb.create_sheet("7. Engineering Plan by Role", index=idx)
for col, w in {"A": 26, "B": 52, "C": 46, "D": 40}.items():
    ws.column_dimensions[col].width = w

ws.merge_cells("A1:D1"); ws.merge_cells("A2:D2")
t = ws.cell(1, 1, "Engineering Plan by Role — What Each Developer Does, Day to Day")
t.font = F(16, True, "FFFFFF"); t.fill = fill(NAVY); t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 26
s = ws.cell(2, 1, "The concrete, ongoing engineering behind 'why a team'. None of these roles is graphic/visual design — this is software engineering.")
s.font = F(c=GREY); s.alignment = CEN; ws.row_dimensions[2].height = 26

# How the pieces connect (framing band)
ws.merge_cells("A3:D3")
band = ws.cell(3, 1, "HOW THE PIECES CONNECT:  The Frontend (TypeScript / React) calls the Backend APIs  →  the Backend APIs talk to the Database. "
                     "Each developer owns one link in that chain and writes the CODE for it. The Tech Lead decides what is built and reviews it all.")
band.font = F(b=True, c=NAVY); band.fill = fill("DDEBF7"); band.alignment = CEN; band.border = BORDER
ws.row_dimensions[3].height = 42
ws.freeze_panes = "A4"
r = 4

def section(title, intro):
    global r
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(r, 1, title); c.font = F(12, True, BLUE); r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(r, 1, intro); c.font = F(c=GREY); c.alignment = TOP; c.border = BORDER
    ws.row_dimensions[r].height = 34; r += 1
    for i, h in enumerate(["Ongoing job", "What the work actually is (day to day)", "Why it matters to the business", "What it prevents"], start=1):
        cc = ws.cell(r, i, h); cc.font = F(b=True, c="FFFFFF"); cc.fill = fill(BLUE); cc.alignment = CEN; cc.border = BORDER
    ws.row_dimensions[r].height = 28; r += 1

def row(vals):
    global r
    for i, v in enumerate(vals, start=1):
        c = ws.cell(r, i, v); c.font = F(b=(i == 1)); c.alignment = TOP; c.border = BORDER
    r += 1

section("FRONTEND DEVELOPER",
        "Writes the TypeScript / React code that runs in the guest's and hotelier's browser and TALKS TO THE BACKEND APIs. "
        "This is engineering — typed code, data handling, performance — not visual design. Builds to the Tech Lead's design and reports to him.")
row(["Call the backend APIs (TypeScript)",
     "Write the typed TypeScript code that calls each backend endpoint: send the correct request, read the response, and keep the frontend's data types matching the backend's contract so data never arrives in the wrong shape.",
     "The screens show exactly the data the backend returns — bookings, rates, availability — correctly and reliably.",
     "Broken pages, wrong data on screen, and type mismatches that crash a component."])
row(["Manage state & data flow",
     "Decide how fetched data moves through the app — caching with React Query, re-fetching on focus, and invalidating the cache after a booking/edit so the screen updates.",
     "Hoteliers and guests always see fresh, correct data without manual refreshes.",
     "Stale screens (e.g. a sold room still showing as available)."])
row(["Loading, error & retry handling",
     "Give every screen a proper 'loading', 'failed' (with a Retry button) and 'empty' state, and tolerate slow or partial API responses.",
     "A slow network or a failed API call never leaves the guest staring at a blank white page mid-booking.",
     "Blank screens and stuck buttons that make guests abandon a booking."])
row(["Performance engineering",
     "Code-splitting, lazy-loading, bundle-size trimming, image optimisation and virtual scrolling for big lists; measure and fix slow renders (Core Web Vitals). (Build/runtime engineering, not graphic design.)",
     "Pages load in ~1-2 seconds even for large hotels — and faster pages convert more guests.",
     "Slow pages that lose bookings, and browsers freezing on thousands of rows."])
row(["Concurrency & input safety",
     "Disable buttons on submit, debounce rapid clicks, and validate inputs in the browser before they reach the API.",
     "No accidental duplicate bookings or double payments on slow networks.",
     "Two bookings / two charges from one impatient double-click."])
row(["Real-time UI",
     "Wire the live WebSocket / SSE streams from the backend into the screens so a new booking appears instantly.",
     "The product feels live; staff react to bookings without refreshing.",
     "Missed or stale information on the dashboard."])

section("BACKEND DEVELOPER",
        "Writes the API code that the frontend calls, and that in turn TALKS TO THE DATABASE. Owns booking logic, payments, security, speed and cost. "
        "Builds to the Tech Lead's design and reports to him.")
row(["Design & build the APIs",
     "Build the endpoints the frontend calls: validate the request, run the business rules, and return the correct JSON shape (the 'contract' the frontend depends on).",
     "Every screen has a reliable, correct API behind it.",
     "Broken or inconsistent responses that break the frontend."])
row(["Talk to the database correctly",
     "Write the queries each API uses to read/write data — always scoped to the caller's hotel, batched (no N+1), and paginated.",
     "Data is correct, isolated per hotel, and fast to fetch.",
     "Cross-tenant leaks, slow queries, and out-of-memory crashes on big hotels."])
row(["API optimisation",
     "Measure each endpoint's response time, cache hot reads in Redis, and push slow work (emails, channel sync) into background jobs.",
     "Responses stay in milliseconds and the server bill stays low — speed is revenue.",
     "Timeouts and a database overloaded by repeated heavy queries."])
row(["Payments & third-party integrations",
     "Razorpay / JioPay, WhatsApp, email: verify signatures, handle webhooks and retries, and ADAPT when a provider changes its API.",
     "Money is handled correctly and integrations keep working as vendors change.",
     "Stuck payments, double charges, and notifications silently breaking."])
row(["Security & continuous bug-hunting",
     "Add the role gate + tenant filter + server-side amount/id verification to every route; read every change for the silent failure modes in Sheet 8.",
     "New features ship without opening data-isolation, tampering or crash holes.",
     "Breaches, price/role tampering, and silent server crashes."])
row(["Cost control & observability",
     "Cap LLM tokens, enforce per-hotel quotas, bound rows/work; emit structured logs (no PII) and error alerts.",
     "Predictable AI/infra cost and problems seen before customers complain.",
     "Surprise bills and outages discovered only via angry hoteliers."])

section("DATABASE DEVELOPER",
        "Owns the database the backend talks to — speed, schema, safety and the guarantee that data flows out smoothly at any size. "
        "Implements the data design the Tech Lead approves and reports to him.")
row(["Schema & migrations",
     "Design the tables and run every schema/index change through one safe, repeatable process (Alembic) instead of ad-hoc startup scripts.",
     "Production never silently misses an index or a column.",
     "Half-applied schemas and 'works locally, crashes in production' bugs."])
row(["Indexing & query plans",
     "Add and tune indexes on the hottest columns (hotel_id, dates, status); read query plans and fix the slow ones.",
     "The busiest queries stay in milliseconds as data grows.",
     "Searches and reports getting slower for everyone over time."])
row(["Partitioning & archiving",
     "Move old bookings/analytics to separate tables; partition large tables by date/hotel.",
     "Day-to-day queries only touch small, recent data — so they stay fast.",
     "Dashboards taking 15-20s to open once a hotel has years of data."])
row(["Connection pooling & capacity",
     "Add a pooler (PgBouncer) and tune pool-size x workers x replicas against the Supabase limit; plan capacity for more hotels.",
     "50-100 hotels can onboard without the database running out of connections.",
     "All tenants slowing down at once when connections run out."])
row(["Backups & restore drills",
     "Run real restore drills, set up point-in-time recovery, and write a disaster-recovery runbook.",
     "Backups are proven to work before we ever need them.",
     "Discovering a backup is unusable on the worst possible day."])
row(["Isolation on every change",
     "Review every new table for Row Level Security and per-hotel isolation.",
     "The database itself keeps refusing to mix one hotel's data with another's.",
     "A schema change quietly opening a cross-tenant leak."])

section("TECH LEAD",
        "Accountable for the WHOLE software — every update, bug and crash. Converts product/stakeholder requirements into an architecture and tasks, "
        "assigns them, reviews every change, and decides what is safe to ship. The three developers above execute his direction and report to him.")
row(["Architecture & decisions",
     "Turn each feature request into a design and an approach; decide what is built, in what order, and how it is coded.",
     "Effort goes to the highest-value work, built the right way the first time.",
     "Rework, wrong-priority work, and inconsistent code."])
row(["Assign & supervise tasks",
     "Break the work into specific tasks for the backend, frontend and database developers and track their progress.",
     "Everyone knows exactly what to build; nothing falls through the cracks.",
     "Developers building in different directions."])
row(["Review gate (production sign-off)",
     "Read every change. By experience, decide whether it is safe to go to production — nothing ships without his approval.",
     "The single most effective catch for the silent bugs in Sheet 8.",
     "A quiet bug reaching production with no experienced second pair of eyes."])
row(["Release & incident response",
     "Own the CI/CD pipeline and the deploy process, and lead the response when something breaks.",
     "Deploys are predictable and incidents are handled calmly.",
     "Risky deploys and slow, chaotic recovery."])
row(["Accountability & security posture",
     "Carry responsibility for every update, bug and crash; schedule pen-tests and drive audit follow-ups.",
     "One clear owner of quality and safety as the product grows.",
     "Diffused responsibility and security drifting as the surface grows."])

# ===========================================================================
# Closing note at the end of Sheet 5 (Future Risks)
# ===========================================================================
ws5 = wb["5. Future Risks"]
r = ws5.max_row + 2
ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
c = ws5.cell(r, 1, "A REALISTIC CLOSING NOTE — WHY THIS WORK NEVER FULLY 'ENDS' (AND WHY A TEAM IS NEEDED)")
c.font = F(12, True, BLUE); r += 1
lines = [
 "In software development, nothing is a permanent solution. The product is healthy today, but it lives in a moving world:",
 "New updates will keep coming — and any update can itself arrive with new bugs.",
 "New features must keep being added to Staybooker to compete — as competitors release features in their software, we have to as well, and every new feature is new surface area where issues can appear.",
 "Third-party APIs change — payment gateways like Razorpay / JioPay update their APIs; if our system does not capture and adapt to those changes in time, errors or bugs can appear.",
 "Infrastructure can be strained — our cloud host (Railway) can be overloaded during traffic spikes.",
 "Real usage is unpredictable — the scenarios in this sheet are informed PREDICTIONS, not certainties. For example: even with 10 hotels, each taking ~15-20 bookings a day (~200 bookings daily), plus 50-100 visitors/guests taking actions on the site at any moment, we cannot predict every combination of what they will do.",
 "So when something does break, there must be a team in place to catch and resolve it BEFORE it grows into a high-priority, customer-facing incident. That is the real reason for the hiring in Sheet 6 — not because the product is broken, but because keeping it healthy is continuous work.",
]
for i, line in enumerate(lines):
    ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    prefix = "" if i == 0 or i == len(lines) - 1 else "•  "
    cc = ws5.cell(r, 1, prefix + line)
    cc.font = F(b=(i == 0 or i == len(lines) - 1)); cc.alignment = TOP; cc.border = BORDER; cc.fill = fill(GREEN)
    ws5.row_dimensions[r].height = 30 if i not in (0, len(lines) - 1) else 42
    r += 1

wb.save(PATH)
print("Saved. Sheet7 rows:", wb["7. Engineering Plan by Role"].max_row, "| Sheet5 rows:", wb["5. Future Risks"].max_row)
