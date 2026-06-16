"""
Stakeholder-grade refresh of the report, mined from the full git history (890 commits).
- Re-title for ALL stakeholders (not just CEO).
- Problems Solved : add a banner + 16 earlier foundational fixes (S-59..S-74).
- Features Built  : add a banner + 14 earlier platform features (#24..#37).
- Work In Progress: correct statuses (load test, refund endpoint, Alembic already exist).
- Future Risks    : add F-14..F-27 (one server, many hotels, many guests).
- Team & Hiring   : strengthen the Tech Lead as the accountable owner + chain-of-command block.
- Engineering Plan by Role : reflect the command structure.
- Executive Summary: refresh counts (74 issues / 37 features).
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PATH = "docs/CEO_Report_Staybooker_Jun2026.xlsx"
NAVY, BLUE, GREY = "0B3050", "1F4E79", "555555"
RED, ORANGE, YELLOW, GREEN, BANNER = "FFC7CE", "FFE3C0", "FFF2CC", "E2EFDA", "DDEBF7"
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP = Alignment(wrap_text=True, vertical="top")
CEN = Alignment(wrap_text=True, vertical="center")
SEV = {"Critical": RED, "High": ORANGE, "Medium": YELLOW, "Low": GREEN}

def F(sz=10, b=False, c=None): return Font(size=sz, bold=b, color=c)
def fill(c): return PatternFill("solid", fgColor=c)

wb = load_workbook(PATH)

def find_row(ws, col, value):
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, col).value).strip() == value:
            return r
    return None

def append(ws, vals, sev_col=None):
    r = ws.max_row + 1
    for i, v in enumerate(vals, start=1):
        c = ws.cell(r, i, v); c.font = F(); c.alignment = TOP; c.border = BORDER
        if sev_col and i == sev_col and v in SEV:
            c.fill = fill(SEV[v]); c.alignment = CEN
    return r

def banner(ws, text, span):
    r = ws.max_row + 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    c = ws.cell(r, 1, text); c.font = F(b=True, c=NAVY); c.alignment = CEN
    c.fill = fill(BANNER); c.border = BORDER
    ws.row_dimensions[r].height = 22
    return r

def setc(ws, r, col, val):
    c = ws.cell(r, col); c.value = val; c.font = F(); c.alignment = TOP; c.border = BORDER

# ===========================================================================
# SHEET 1 — re-title for stakeholders + refresh counts
# ===========================================================================
ws = wb["1. Executive Summary"]
ws.cell(1, 1).value = "Staybooker Booking Engine — Stakeholder Report (June 2026)"
ws.cell(2, 1).value = ("Prepared for the CEO, founders, investors and stakeholders. What has been solved, what is in progress, "
                       "what risks are coming as we grow, how the product stays fast and crash-proof at scale, and the team needed to sustain it.")
def repl(topic, detail):
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value).strip() == topic:
            ws.cell(r, 2).value = detail; return
repl("Sheet 2 — Problems Solved",
     "74 issues fixed so far (6 Critical, 38 High, 30 Medium) across the project's history. Each issue shows what would have happened to the business if it had reached production.")
repl("Sheet 3 — Features Built",
     "37 major product features delivered alongside the bug fixing — multi-property chains, loyalty, the Revenue Engine, OTA-style booking widgets, a GST/tax engine, WhatsApp & Google AI agents, a full Super Admin platform console, subscriptions & feature-gating, and more.")
repl("Issues fixed till date",
     "74 (source: full git history of 890 commits + work_log_tracker.csv, up to 16-Jun-2026; merged across PRs up to #90)")
repl("Automated tests added",
     "Close to 300 automated test cases (security, payments, booking, roles, rate limits, Redis failover, chain isolation, promotions) that run on every code change (CI/CD), with dependency vulnerability (CVE) scanning.")

# ===========================================================================
# SHEET 2 — Problems Solved : earlier foundational fixes
# ===========================================================================
ws = wb["2. Problems Solved"]
banner(ws, "EARLIER FOUNDATIONAL FIXES (May – early June 2026) — security, crashes and correctness fixed before the June audit", 7)
earlier = [
 ("S-59","The public booking page crashed for guests (blank/white page) more than once","React runtime errors (#300 and a hooks-rule violation in a booking widget) crashed the page render","Guests opening the booking page could see a blank or broken page and leave — lost bookings with no error trail.","Fixed the React errors and hook usage; added guards so one widget can't crash the whole page.","High","02-03 Jun"),
 ("S-60","The backend could crash on startup","Superadmin import errors, a sync/async mismatch and an undefined router crashed the server on boot","A bad deploy would take the whole API down — every hotel offline until someone noticed.","Fixed the imports, converted to async correctly, and removed the undefined router.","High","May 12-Jun 01"),
 ("S-61","The rooms page crashed with a 'Form is not defined' error","A circular dependency loaded a dialog before its form was ready","Hoteliers couldn't manage rooms — a core daily task was broken.","Lazy-loaded the dialogs to break the circular import.","Medium","14-May"),
 ("S-62","The public booking page crashed on newly created rooms","New rooms weren't handled in the availability path and the rate cache wasn't invalidated","Brand-new inventory couldn't be booked — directly blocking sales of the newest rooms.","Fixed the availability handling and added real-time rate-cache invalidation.","High","23-May"),
 ("S-63","The guest chatbot crashed on availability follow-up questions","The Llama model returned a tool_use_failed error that wasn't handled","A premium guest feature broke mid-conversation, frustrating guests.","Added handling for the failure path so the chat recovers gracefully.","Medium","03-Jun"),
 ("S-64","Database migrations could fail on deploy (divergent Alembic heads)","Two migration branches diverged and some migrations weren't idempotent","A deploy could half-apply the schema and break production.","Merged the migration heads and made migrations safe to re-run.","High","03-Jun"),
 ("S-65","Switching between properties could briefly show another property's data","Browser-side caching and a backend cache-key bug served stale, cross-property data","A hotelier could momentarily see the wrong hotel's numbers — a trust and isolation concern.","Disabled unsafe API caching, fixed the cache key, and moved to instant SPA navigation.","High","04-05 Jun"),
 ("S-66","Cross-tenant data-isolation gaps in several endpoints (IDOR)","Some endpoints didn't fully scope queries to the caller's hotel/chain (TEN-02/04/05)","One hotel could potentially reach another's data — a serious multi-tenant breach risk.","Tightened scoping on every affected endpoint to the caller's own tenant.","High","05-Jun"),
 ("S-67","Role-based access was defined but not actually enforced on some routes","RBAC existed in code but several routes never checked it (AZ-01/02/03)","Staff-level accounts could reach owner-level actions — privilege misuse.","Enforced the role checks on every affected route.","High","05-Jun"),
 ("S-68","AI spend on public/agent paths was unbounded","Some LLM calls had no token or call ceiling (AI-01/02/04)","A traffic spike or abuse could run the AI bill up sharply.","Added token floors, tool-call limits and history caps across the AI paths.","High","05-Jun"),
 ("S-69","Reports and analytics could show wrong numbers","Booking-source attribution and several analytics calculations were incorrect","Hoteliers would make decisions on wrong data and lose trust in the product.","Corrected the attribution and the analytics math; added tests.","Medium","04-Jun"),
 ("S-70","Dashboards were slow and overloaded the database (N+1 queries)","Pages fired many small queries instead of one batched query; hotel info wasn't cached","As data grew, dashboards would slow for everyone and strain the database.","Batched the queries, added caching and parallelised integration calls.","Medium","02-04 Jun"),
 ("S-71","Responses had no security headers","Missing HSTS, CSP, X-Frame-Options and X-Content-Type-Options","The app was more exposed to clickjacking, content-sniffing and mixed-content attacks.","Added enterprise security headers across all responses.","Medium","16-May"),
 ("S-72","A null widget-config value could crash the embedded booking widget","Missing null-safety on widget configuration and chat","The widget could break on a hotel's own website, in front of their guests.","Hardened the config handling and chat resilience.","Medium","04-May"),
 ("S-73","The AI usage dashboard was stuck on 'Loading…' / 'Couldn't load usage'","Subscription AI-limit columns were missing and usage lived only in Redis","Hoteliers couldn't see what they were paying for, and data vanished on a Redis blip.","Added the missing columns and made AI usage durable in Postgres.","Medium","12-Jun"),
 ("S-74","Guests couldn't reset their password reliably","The reset flow didn't fully support PKCE/implicit hand-off and recovery redirects","Locked-out users couldn't get back in — support load and churn.","Completed the reset flow (PKCE/implicit, recovery redirect, in-profile change).","Medium","05-Jun"),
]
for row in earlier:
    append(ws, row, sev_col=6)

# ===========================================================================
# SHEET 3 — Features Built : earlier platform features
# ===========================================================================
ws = wb["3. Features Built"]
banner(ws, "EARLIER PLATFORM FEATURES (built across May – early June 2026)", 5)
features = [
 (24,"GST / tax engine","Configurable tax slabs (inclusive & exclusive), itemized tax breakdowns on every booking and invoice, and guest-facing price/tax disclosures.","Correct, compliant billing for Indian hotels — essential for trust and GST filing.","Tax is validated on save and computed server-side; guarded against NaN/zero-division."),
 (25,"WhatsApp AI Agent","Per-hotel WhatsApp integration that sends booking confirmations and answers guests with AI, metered against subscription credits.","Guests get instant answers and confirmations on the channel they actually use.","Webhook payloads HMAC-verified; credits deducted per message; keys resolved from the encrypted Vault."),
 (26,"Google Reviews AI Reply Studio","Per-hotel Google OAuth connect, fetch reviews, AI-drafted replies, and one-click posting back to the Google Business Profile.","Saves staff hours and improves the hotel's public reputation — a premium upsell.","Per-hotel OAuth tokens stored securely; scoped to the connected account."),
 (27,"Google Hotel Ads integration","Feeds and UI to list a hotel's rates and availability on Google Hotel Ads.","A new demand channel for direct bookings.","Feature-gated by plan; per-hotel scoped."),
 (28,"Multi-tenant email system","Custom email via SMTP or Brevo, with per-hotel sender, CC, signature and branded templates.","Confirmations and communications come from the hotel's own brand, not a generic system.","Per-hotel credentials encrypted; senders fail-safe so a mail error never breaks a booking."),
 (29,"Super Admin platform console","A full operator console — KYC, commissions, payouts, support tickets, platform settings, MRR & plan breakdowns, system diagnostics, and secure impersonation with an exit banner and audit trail.","Lets us run Staybooker as a business: onboard, bill, support and oversee every hotel.","Superadmin-gated; impersonation tokens expire; every privileged action is audit-logged."),
 (30,"Subscriptions & feature-gating","Subscription plans with a per-plan feature-allocation matrix, automatic feature sync, and premium locks across the product.","Turns features into packaged, sellable tiers — the core of the revenue model.","Plan checks enforced server-side; writes to subscription state are locked."),
 (31,"Customizable role permissions","A per-hotel role-permission matrix (OWNER/MANAGER/STAFF and sub-roles) with a complete security audit trail.","Hotels control exactly what each employee can do — important for larger properties.","Permissions enforced on the backend, not just the menu; every change is logged."),
 (32,"Configurable AI stack","AI agents on the Agno framework with per-hotel model choice (Llama, Google Gemini, DeepSeek), token limits, and a durable per-agent usage dashboard.","Right-sized AI cost and quality per hotel, with full visibility into spend.","Cheap model by default; token/tool-call limits; usage durably tracked in Postgres."),
 (33,"Experiences & Activities upsell","Add-ons, services and amenities a guest can add at checkout (e.g. activities, transfers), managed by the hotel.","Higher average booking value — incremental revenue on every stay.","Prices validated server-side; tenant-scoped."),
 (34,"Configurable cancellation & refund policies","Instant-vs-request cancellation modes (set by super admin) and room-level cancellation/refund overrides.","Hotels match the platform to their real policies — fewer disputes.","Mode enforced server-side; refunds run through the verified payment path."),
 (35,"Booking source tracking & leads","Attribution of where each booking came from, AI-captured leads with filters, and CSV export of bookings and leads.","Hotels see which channels work and can follow up on leads — more direct bookings.","Tenant-scoped; exports limited to the hotel's own data."),
 (36,"Embeddable widget suite","Multiple booking-widget layouts (Booking.com-style, STAAH-style, capsule, glassmorphism), theme-matched, with live per-date pricing and dual hotel + chain widgets.","Hotels add a professional, commission-free booking engine to their own website.","Host-page-safe loader; hardened postMessage; live prices confirmed server-side."),
 (37,"Public presence & compliance","A premium landing page with smart subdomain routing, customizable hotel URL slugs, and the legal/compliance pages (refund, cookie, contact, Meta-verification).","A credible public face and the compliance pages required to operate and integrate.","Slug uniqueness validated; compliance pages public by design."),
]
for row in features:
    append(ws, row)

# ===========================================================================
# SHEET 4 — Work In Progress : correct statuses to reality
# ===========================================================================
ws = wb["4. Work In Progress"]
def wip(pid, why, status):
    r = find_row(ws, 1, pid)
    if r:
        setc(ws, r, 3, why); setc(ws, r, 4, status)
wip("P-01",
    "A Razorpay refund endpoint, idempotency keys and webhook signature verification are already built. The remaining pieces are the FULLY-AUTOMATIC refund flow and a daily reconciliation job (matching the Razorpay ledger to our bookings).",
    "In Progress (foundations done)")
wip("P-02",
    "A load-test harness and scale documentation (SCALE_TUNING.md) already exist, and database row-locks protect against double-selling a room. The remaining work is running load tests at realistic scale and tuning the connection pool / auto-scaling from the results.",
    "In Progress (harness ready)")
wip("P-05",
    "Alembic is set up (divergent heads merged, migrations made idempotent), but several newer features still create tables via create_all/init_db at startup. The remaining work is to bring every schema change under one consistent Alembic process.",
    "In Progress (Alembic exists)")

# ===========================================================================
# SHEET 5 — Future Risks : one server, many hotels, many guests
# ===========================================================================
ws = wb["5. Future Risks"]
# cols: ID|Scenario|What Can Happen|Why It Will Happen|Protection Already In Place|What Still Needs To Be Built|Severity|Owner
banner(ws, "MORE SCENARIOS AS WE SCALE — many hotels and their guests all share the same servers", 8)
risks = [
 ("F-14","One busy hotel's traffic or heavy reports slow down every other hotel ('noisy neighbour')","All hotels share the same servers and database; one hotel running a huge report or a sale can starve the rest","Shared multi-tenant infrastructure with finite CPU, connections and memory","Pagination and row caps everywhere; per-hotel AI quotas; Redis caching; env-driven pools","Per-tenant rate/usage limits, query budgets, and isolating heavy reports onto background workers/replicas","High","Backend + Database + DevOps"),
 ("F-15","A popular cached page expires and hundreds of requests hit the database at once (cache stampede)","A sudden database spike can slow or stall the site exactly when traffic is highest","When one cache key expires under heavy load, every request rebuilds it simultaneously","Redis caching with an in-memory fallback layer","Stale-while-revalidate, request coalescing (single-flight) and jittered cache TTLs","High","Backend"),
 ("F-16","Two staff or two guests update the same inventory at the same moment and the database deadlocks","Some requests fail with errors during the busiest booking moments","Concurrent writes to the same rows across many hotels can deadlock in Postgres","Row-level locks (SELECT FOR UPDATE) on inventory and redemption","Consistent lock ordering, retry-on-deadlock, and shorter transactions","Medium","Database"),
 ("F-17","Email/WhatsApp providers throttle or reject our messages and guests stop getting confirmations","Guests miss confirmations and OTAs aren't synced — more support calls, double bookings","Brevo/Meta enforce sending limits and deliverability rules that tighten as volume grows","Senders are fail-safe (never crash the booking); per-hotel credentials","A queue with retry/back-off, deliverability monitoring, and provider fail-over","High","Backend"),
 ("F-18","A hotel's WhatsApp/Meta token expires or a template is rejected and notifications silently stop","That hotel's WhatsApp confirmations quietly stop working","Meta tokens expire and message templates need approval; both change over time","Per-hotel config; webhook signature verification","Token-expiry alerts, template-status checks, and a clear 'reconnect' prompt","Medium","Backend"),
 ("F-19","Google review/Business-Profile OAuth tokens expire and the review features silently stop","The reviews studio stops fetching or posting for that hotel","OAuth refresh tokens can be revoked or expire","Per-hotel OAuth tokens stored securely","Automatic token refresh, expiry alerts, and a guided reconnect flow","Medium","Backend"),
 ("F-20","A Razorpay webhook is replayed or arrives out of order and a payment is processed twice","A guest could be double-charged or a booking double-confirmed","Webhooks can be retried, duplicated or delayed by the network","Idempotency keys; webhook signature verification; status guard on confirmed bookings","Event de-duplication by event id and an ordered, idempotent webhook processor","High","Backend"),
 ("F-21","Hotels in different regions hit date/timezone bugs (check-in/out or reports off by a day)","Wrong nights charged, wrong availability, or reports that don't match the hotel's day","Dates are computed in multiple layers; timezones differ per hotel","Server-side price and date validation","A single, explicit timezone strategy per hotel and timezone tests","Medium","Backend + Frontend"),
 ("F-22","Stored data (guest images, invoices, backups) grows unbounded as hotels and stays pile up","Storage costs climb and large objects slow pages and backups","Every booking and hotel adds images and documents that are never cleaned up","Randomized filenames; size caps on uploads","Object storage on a CDN, lifecycle/archival rules, and an image-processing pipeline","Medium","Database + DevOps"),
 ("F-23","A bug loop or an attack floods the logs/monitoring and the observability bill explodes","Monitoring costs spike and real errors get buried in noise","High traffic plus an error loop can generate enormous log volume","Sentry sampling reduced in production; no PII in logs","Log rate-limiting, error de-duplication, and budget alerts on observability spend","Medium","Backend + DevOps"),
 ("F-24","A new hotel onboards with a bad bulk import (rooms/rates/bookings) and pages break","The new hotel's dashboard or booking page errors out on day one","Imported data may have missing fields, wrong types or duplicates","Server-side validation on the main write paths","A validated import pipeline with a dry-run preview and clear error reporting","Medium","Backend + Frontend"),
 ("F-25","An AI provider has an outage or deprecates a model and the chatbots/agents stop or misbehave","Guest and hotelier AI features go down or change behaviour suddenly","We depend on third-party LLMs whose availability and models change","Cheap model by default; per-hotel model selection; usage tracking","Multi-provider fail-over, model pinning, and a graceful 'AI unavailable' fallback","Medium","Backend"),
 ("F-26","A guest exercises their DPDP/GDPR right to be forgotten and we must delete their data correctly","A mishandled deletion is a compliance violation; an over-broad one corrupts records","Privacy law requires verifiable deletion across many tables and many hotels","Per-hotel data isolation; encrypted PII","A documented data-deletion/anonymisation workflow and an audit record of each request","Medium","Backend + Database"),
 ("F-27","A vendor key leaks and must be rotated across all hotels without downtime","A compromised key must be replaced fast, for every affected hotel, with no outage","Keys can leak through logs, laptops or third parties; we hold many per hotel","All secrets encrypted in the Vault; never returned in plain text; masked in APIs","A documented key-rotation runbook and per-hotel re-key tooling","High","Backend + Tech Lead"),
]
for row in risks:
    append(ws, row, sev_col=7)

# ===========================================================================
# SHEET 6 — Team & Hiring : strengthen Tech Lead + chain of command
# ===========================================================================
ws = wb["6. Team & Hiring"]
# Data-journey step 4 (row 16)
setc(ws, 16, 1, "4. Someone decides WHAT to build, HOW to build it, reviews it, and signs off before release")
setc(ws, 16, 3, ("TECH LEAD — the decision-maker and the final gate. Requirements from the product manager / stakeholders come to him; he designs the architecture, "
                 "breaks it into tasks for each developer, reviews every change, and by his experience decides whether it is safe to ship to production. The developers below "
                 "build to his direction and report to him. He is accountable for every software update, every bug and every crash."))
# Dev rows in the data journey - add that they execute the Tech Lead's plan
for rr in (13, 14, 15):
    cur = ws.cell(rr, 3).value
    if cur and "to the Tech Lead's" not in cur:
        setc(ws, rr, 3, cur + " Builds to the Tech Lead's design and reports progress to him.")
# Role summary (row 21 Tech Lead; 22-24 devs)
setc(ws, 21, 2, ("The single person accountable for the whole software — every update, bug and crash. He decides what gets built, when and how it is coded, "
                 "reviews and approves every change before it can reach production, and the developers execute his plan and report to him."))
setc(ws, 21, 3, ("Turns feature requests from the product manager / stakeholders into an architecture and a set of tasks; assigns each task to the right developer; "
                 "reviews ALL code and, by experience, decides what is safe to ship to production; owns the engineering rules, the review checklist and the CI/CD gate; "
                 "carries responsibility when anything breaks. (In this project: set the fix order across 70+ issues money-first, wrote the rules the team follows, and stood up the test/CI gate.)"))
for rr, extra in ((22, " Builds what the Tech Lead designs — the 'what' and 'why' come from him; reports progress back to him."),
                  (23, " Builds the screens to the Tech Lead's design and direction; reports to him."),
                  (24, " Implements the data design the Tech Lead approves; reports to him.")):
    cur = ws.cell(rr, 2).value
    if cur and "Tech Lead" not in cur:
        setc(ws, rr, 2, cur + extra)
# Hiring order Tech Lead (row 32)
setc(ws, 32, 3, ("Direction and review from day one. He is the owner of quality and safety — nothing reaches production without his approval — so the work of every other "
                 "hire stays consistent and safe. Should be a senior, hands-on engineer who can also code."))
setc(ws, 32, 4, "Set the technical roadmap and architecture; define the code-review and release process; assign and supervise all developer tasks; own incident response.")
# What each hire returns Tech Lead (row 42)
setc(ws, 42, 2, "No single owner of quality; code reaching production with no experienced second pair of eyes; no clear accountability when something breaks.")
setc(ws, 42, 3, ("One accountable owner for the entire software — every update, bug and crash. He designs each feature, assigns the tasks, and personally approves what ships, "
                 "so quality and safety do not depend on luck."))
setc(ws, 42, 4, "Developers build in different directions, reviews get skipped, and a single silent bug (see the Silent Bugs sheet) can reach production unchecked.")

# Append a clear chain-of-command block
r = ws.max_row + 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws.cell(r, 1, "HOW THE TEAM WORKS — THE CHAIN OF COMMAND"); c.font = F(b=True, sz=12, c=BLUE)
r += 1
header = ["Step", "Who", "What happens"]
for i, h in enumerate(header, start=1):
    cc = ws.cell(r, i, h); cc.font = F(b=True, c="FFFFFF"); cc.fill = fill(BLUE); cc.alignment = CEN; cc.border = BORDER
ws.row_dimensions[r].height = 22
chain = [
 ("1", "Product Manager / Stakeholders", "Decide WHAT the business needs — new features, priorities, customer requests — and hand them to the Tech Lead."),
 ("2", "Tech Lead", "Architects the design and the approach, decides HOW and WHEN it is built, and breaks it into specific tasks for each developer."),
 ("3", "Backend / Frontend / Database Developers", "Build their assigned tasks to the Tech Lead's design. They focus on writing the code; the 'what' and 'why' come from the Tech Lead."),
 ("4", "Tech Lead (review gate)", "Reviews every change. By his experience he decides whether it is safe to go to production — nothing ships without his approval."),
 ("5", "Tech Lead (accountability)", "Owns the result: any update, bug or crash is his responsibility. The developers report to him; he reports to the business."),
]
for step, who, what in chain:
    r += 1
    for i, v in enumerate((step, who, what), start=1):
        cc = ws.cell(r, i, v); cc.font = F(b=(i <= 2)); cc.alignment = TOP; cc.border = BORDER

# ===========================================================================
# SHEET 7 — Engineering Plan by Role : reflect command structure
# ===========================================================================
ws = wb["7. Engineering Plan by Role"]
for r in range(1, ws.max_row + 1):
    a = str(ws.cell(r, 1).value or "")
    if a == "TECH LEAD":
        # the intro line is the next row, merged in col 1
        intro_r = r + 1
        ws.cell(intro_r, 1).value = ("Accountable for the WHOLE software — every update, bug and crash. He converts product/stakeholder requirements into an architecture and tasks, "
                                     "assigns them to the developers, reviews every change, and decides what is safe to ship to production. The developers execute his direction and report to him.")
    if a in ("BACKEND DEVELOPER", "FRONTEND DEVELOPER", "DATABASE DEVELOPER"):
        intro_r = r + 1
        cur = ws.cell(intro_r, 1).value or ""
        if "Tech Lead" not in cur:
            ws.cell(intro_r, 1).value = cur.rstrip() + " Works to the Tech Lead's design and direction, and reports back to him."

wb.save(PATH)
print("Saved. Sheet sizes:")
for s in wb.worksheets:
    print(f"  {s.title}: rows={s.max_row}")
