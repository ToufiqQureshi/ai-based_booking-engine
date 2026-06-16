"""
Refresh the CEO report so every sheet reflects the latest work
(work_log_tracker.csv 14-16 Jun + merged PRs #80/#82/#83/#84/#85/#86/#88/#89/#90).

- Sheet 2 Problems Solved : append S-46..S-58 (newly solved), mark as fixed.
- Sheet 3 Features Built   : append the Revenue Engine, chain features, widget, etc.
- Sheet 4 Work In Progress : update each item's status to current reality.
- Sheet 5 Future Risks     : refresh 'Protection Already In Place' where new protection landed.
- Sheet 1 Executive Summary: refresh the numbers and add a 'what's new' block.
- Sheet 6 Team & Hiring     : refresh issue counts and examples.
- Sheet 8 Silent Bugs       : cite the new real crash fixes.

Run from repo root:  python scripts/update_ceo_report.py
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PATH = "docs/CEO_Report_Staybooker_Jun2026.xlsx"
NAVY, BLUE, GREY = "0B3050", "1F4E79", "555555"
RED, ORANGE, YELLOW, GREEN = "FFC7CE", "FFE3C0", "FFF2CC", "E2EFDA"
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPWRAP = Alignment(wrap_text=True, vertical="top")
CENTERWRAP = Alignment(wrap_text=True, vertical="center")
SEV = {"Critical": RED, "High": ORANGE, "Medium": YELLOW, "Low": GREEN, "Done": GREEN}

def body(sz=10, bold=False, color=None):
    return Font(size=sz, bold=bold, color=color)
def fill(c):
    return PatternFill("solid", fgColor=c)

wb = load_workbook(PATH)

def find_row(ws, col, value):
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, col).value).strip() == value:
            return r
    return None

def append_row(ws, values, sev_col=None):
    r = ws.max_row + 1
    for i, v in enumerate(values, start=1):
        c = ws.cell(r, i, v)
        c.font = body()
        c.alignment = TOPWRAP
        c.border = BORDER
        if sev_col and i == sev_col and v in SEV:
            c.fill = fill(SEV[v])
            c.alignment = CENTERWRAP
    return r

def set_cell(ws, r, col, value, keepstyle=True):
    c = ws.cell(r, col)
    c.value = value
    if not keepstyle:
        c.font = body(); c.alignment = TOPWRAP; c.border = BORDER
    return c

# ---------------------------------------------------------------------------
# SHEET 2 — Problems Solved : append S-46..S-58
# ---------------------------------------------------------------------------
ws = wb["2. Problems Solved"]
new_problems = [
    ("S-46", "The loyalty check could crash the public booking page for some guests",
     "The loyalty lookup used scalar_one_or_none(), which throws when a guest matches more than one loyalty row",
     "A guest pressing 'check loyalty' could hit a 500 error mid-booking — a silent server crash that only appears for certain data, dropping bookings with no error anyone could easily reproduce.",
     "Switched to .scalars().first() with a clear preferred-row ordering; added a regression test for the exact path.", "High", "14-Jun"),
    ("S-47", "Creating or cancelling a public booking could crash with a 500",
     "The code read super_admin.phone (a field that does not exist) and WhatsApp settings that were never defined",
     "The core revenue action of the whole product — creating/cancelling a booking — could fail server-side, silently, for everyone.",
     "Admin alert number moved to server config (env); WhatsApp service aligned to the real settings; two unused admin lookups removed (also fewer DB queries per booking).", "High", "15-Jun"),
    ("S-48", "The whole app could load as a blank white screen after a deploy",
     "A production build optimisation (Vite manualChunks) broke React's context loading and crashed the app to a blank page",
     "Guests and hoteliers would open the site to a completely blank screen — the worst possible failure, with no error and no way forward.",
     "Reverted the unsafe bundling change and added instant-load placeholders so the first paint is never empty.", "High", "15-Jun"),
    ("S-49", "Abandoned-booking recovery was built but never actually ran",
     "The recovery feature had no scheduled job, so unpaid bookings always aged past the 72h window un-nudged",
     "A revenue-recovery feature we built was silently dead — incremental bookings left on the table every day with no one noticing.",
     "Added an hourly Redis-locked scheduler that runs the sweep on exactly one server and honours each hotel's opt-in.", "Medium", "15-Jun"),
    ("S-50", "The booking widget could run attacker-supplied code on the hotel's site (RCE/XSS)",
     "A 'custom JS' setting was executed via new Function()/eval on the public widget; widget CSS and chat messages were under-validated",
     "A malicious or compromised setting could run arbitrary code in every guest's browser on the hotel's own domain — a severe security hole.",
     "Removed the code-execution sink entirely; the public API no longer returns custom JS; iframe sandboxed; CSS sanitised (strips url()/@import/expression/script); chat postMessage locked to the referrer origin; inactive hotels return 404; guest message length capped.", "Critical", "15-Jun"),
    ("S-51", "Two guests could spend the same promo / loyalty balance at once",
     "Promo-code and loyalty redemption read-then-wrote without a row lock (a race condition)",
     "Under concurrent checkouts the same discount or points balance could be spent twice — direct revenue leakage and wrong loyalty liability.",
     "Redemption now takes SELECT FOR UPDATE row locks; AI lead inputs are length-bounded as well.", "High", "15-Jun"),
    ("S-52", "A newly created property could be saved without its brand (chain) link",
     "The property-create path did not inherit chain_id from the logged-in user",
     "A property could end up unlinked or wrongly associated across tenants — a multi-tenant isolation risk and broken brand dashboards.",
     "chain_id is now inherited from the authenticated user on creation; the superadmin add-property path also sets address fields correctly.", "High", "14-Jun"),
    ("S-53", "Loyalty read endpoints had no role check",
     "GET /loyalty/program and /loyalty/guests were missing the role guard the write endpoints already had",
     "Any staff account could read the loyalty program configuration and the guest list — an authorization gap on read (a trap we watch for everywhere).",
     "require_hotel_role added to both GET endpoints.", "Medium", "14-Jun"),
    ("S-54", "The frontend could not deploy at all (Cloudflare Pages broken)",
     "A stray git submodule reference (decodo-api-docs) broke every Cloudflare Pages build",
     "No frontend deploys could ship — the product could not be updated for customers until it was fixed.",
     "Removed the broken submodule gitlink; deploys restored.", "Medium", "15-Jun"),
    ("S-55", "The frontend production build was silently broken on the branch",
     "A new widget tab imported a UI component (accordion) whose file was never committed — 'vite build' failed with a missing-file error",
     "The next release would have failed to build with no obvious cause — exactly the kind of quiet break that blocks a deploy at the worst time.",
     "Added the missing standard component (the dependency was already present) and verified with a full local build.", "Medium", "16-Jun"),
    ("S-56", "Some database columns existed in code but not in the live database",
     "Migrations for loyalty milestones and competitor-rate columns were missing from the deploy path",
     "Any query touching those columns would crash in production while passing locally — a hidden landmine as features rolled out.",
     "Added the missing migrations so the live schema matches the code.", "High", "15-Jun"),
    ("S-57", "Hotel listings and sister-hotel suggestions always showed a blank city",
     "City was read as a direct attribute, but it is stored inside the address JSON dict",
     "Guests saw blank locations on chain pages and recommendations — it looks broken and quietly hurts trust and conversion.",
     "All three hotel-listing endpoints now read city from the address dict.", "Medium", "14-Jun"),
    ("S-58", "Valid bookings were sometimes rejected over a 1-rupee rounding gap",
     "The client/server price match used a flat +/-5 INR tolerance that failed on percentage-based dynamic prices",
     "Guests with legitimate prices could be blocked at checkout by a tiny rounding difference — silent lost bookings.",
     "Tolerance is now max(Rs.5, 1%) and the calendar pricing note is clearer.", "Medium", "15-Jun"),
]
for row in new_problems:
    append_row(ws, row, sev_col=6)

# ---------------------------------------------------------------------------
# SHEET 3 — Features Built : append #13..#23
# ---------------------------------------------------------------------------
ws = wb["3. Features Built"]
new_features = [
    (13, "Dynamic pricing engine (hotelier-configured)",
     "Hoteliers set rules (occupancy, day-of-week, lead-time) that adjust room prices automatically; a live preview shows the price on each calendar cell, and the rules drive the real booking charge.",
     "A revenue-management feature (higher RevPAR) that competitors charge a premium for — a new selling point and revenue line.",
     "Prices are server-computed with live occupancy; the client's price is verified within a tolerance; rules are role-gated and tenant-scoped."),
    (14, "Abandoned-booking recovery",
     "Automatically nudges guests who started but did not finish a booking, via WhatsApp and email, on an opt-in basis.",
     "Recovers otherwise-lost bookings automatically — direct incremental revenue.",
     "Runs on an hourly Redis-locked scheduler (one server only); per-hotel opt-in; senders never crash the booking flow; no PII logged."),
    (15, "Loyalty points wallet + stay offers",
     "Configurable earn/redeem rates, room-specific night-threshold offers, and a pre-booking nudge ('stay one more night to unlock').",
     "Deeper loyalty and longer stays — more direct, OTA-free revenue.",
     "Redemption value is recomputed server-side from config (never trusts the client); role-gated; brand-isolated."),
    (16, "Chain-wide upsell broadcast",
     "A brand admin pushes one stay-upsell offer to every active property at once; each hotel gets its own row so the guest flow is unchanged.",
     "Brand-level marketing in a single click across all properties.",
     "Chain-admin only; isolated by chain_id; per-hotel rows share a broadcast id."),
    (17, "Seasonal auto-apply promotions",
     "Date-windowed deals that apply automatically with no coupon code (or as classic coupons), at hotel and chain level.",
     "Easier campaigns and higher conversion (no coupon friction).",
     "Discount always server-computed; usage incremented atomically; scoped to hotel or chain; the display banner is read-only."),
    (18, "OTA / Booking.com-style booking widget",
     "A professional, embeddable booking + search widget for hotel websites, with a floating calendar overlay and instant-load placeholders.",
     "A polished direct-booking experience that rivals the OTAs — captures commission-free bookings.",
     "Host-page-safe embed loader; hardened postMessage; iframe sandbox; the code-execution sink was removed (see S-50)."),
    (19, "Brand Console for multi-property owners",
     "A collapsible brand workspace (overview, guest insights, brand promos, loyalty, upsell broadcast) for groups managing many hotels.",
     "Makes the platform genuinely usable for hotel groups, not just single properties.",
     "Gated by chain membership; every section validates the user's chain."),
    (20, "Configurable rate plans & packages",
     "Pure rate plans are separated from packages; packages gain an optional validity window and a 'you save' badge.",
     "Clearer pricing setup for hoteliers and fewer mistakes.",
     "Tenant-scoped; reuses the existing rate endpoints."),
    (21, "Anti-automation booking token (HMAC)",
     "A signed, short-lived token on the public booking path to deter scripted abuse (feature-flagged, off by default).",
     "Protects the booking funnel from bots without slowing real guests.",
     "HMAC-signed, time-bounded and rate-limited; enabled only when needed."),
    (22, "Production observability (Sentry)",
     "Error tracking with user context, a booking-funnel trail and payment-failure tracking.",
     "We see and fix issues before customers report them — higher reliability.",
     "No PII or secrets in events; scoped server-side."),
    (23, "Engineering knowledge graph (graphify)",
     "An AST-built map of the whole codebase (~4,800 nodes) that any engineer — or the AI assistant — consults before changing code.",
     "Faster, safer onboarding and changes as the team grows; fewer regressions.",
     "Local-only, no API cost; rebuilds automatically on commits."),
]
for row in new_features:
    append_row(ws, row)

# ---------------------------------------------------------------------------
# SHEET 4 — Work In Progress : refresh statuses to current reality
# ---------------------------------------------------------------------------
ws = wb["4. Work In Progress"]
# cols: ID | Task | Why It Matters | Status | Risk If Delayed
wip_updates = {
    "P-03": ("Frontend page-by-page error audit (errors + edge cases on every page)",
             "Real progress since the last report: the blank-screen (createContext) crash is fixed (S-48), instant-load placeholders and skeleton price cells are in, and the booking-widget pages were hardened (S-50, plus the earlier S-34/S-36/S-37). The hotelier dashboard pages are the remaining batch.",
             "In Progress"),
    "P-04": ("DDoS / abuse protection layer (WAF)",
             "The IP-spoofing hole is fixed (S-40), key routes are rate-limited (S-42), and a signed anti-automation booking token (HMAC) now guards the booking path. The remaining piece is an edge layer (Cloudflare/WAF) to absorb large-scale attacks before they reach our servers.",
             "In Progress"),
    "P-05": ("Database migration discipline (Alembic in the deploy process)",
             "Several missing migrations were added (S-56) and new features now ship with their own indexes (loyalty offers, broadcast id, promo auto-apply, package validity). The schema is still built by create_all/init_db at startup, so a proper Alembic migration process is the remaining work.",
             "In Progress"),
    "P-06": ("Remaining medium-severity audit findings",
             "Several have been closed since: the widget code-execution sink was removed (S-50), chat postMessage origin locked, and loyalty read endpoints role-gated (S-53). Items still open: the login-token cache window (~10 min), the unsigned OAuth state parameter, and stale analytics cache invalidation.",
             "In Progress"),
}
for pid, (task, why, status) in wip_updates.items():
    r = find_row(ws, 1, pid)
    if r:
        set_cell(ws, r, 2, task)
        set_cell(ws, r, 3, why)
        set_cell(ws, r, 4, status)

# Append a 'recently completed' note row pair under WIP
r = ws.max_row + 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws.cell(r, 1, "RECENTLY COMPLETED (graduated out of 'in progress' since the last report)")
c.font = body(bold=True, color=BLUE)
r += 1
done_note = ("Revenue Engine shipped (dynamic pricing, abandoned-booking recovery, loyalty wallet & upsell hub) - PR #80; "
             "chain features, package improvements & seasonal promotions - PR #88; the recovery scheduler and the booking create/cancel 500 crashes - PR #83; "
             "the Cloudflare blank-screen crash - PR #82; the broken-deploy and missing-migration fixes - PR #84/#85; production observability (Sentry) and the engineering knowledge graph - PR #59/#86. "
             "All are reflected in Sheets 2 and 3.")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws.cell(r, 1, done_note)
c.font = body(); c.alignment = TOPWRAP; c.border = BORDER; c.fill = fill(GREEN)
ws.row_dimensions[r].height = 60

# ---------------------------------------------------------------------------
# SHEET 5 — Future Risks : refresh 'Protection Already In Place' (col F)
# ---------------------------------------------------------------------------
ws = wb["5. Future Risks"]
PROT = 5  # 'Protection Already In Place' column
prot_updates = {
    "F-04": "Redis-backed rate limiting; spoof-proof client IP detection (trusted proxy); per-hotel AI quotas; per-route caps on password change/uploads/reports; limits on streaming connection lifetimes; PLUS a signed anti-automation booking token (HMAC) on the booking path.",
    "F-06": "Backend idempotency ensures the same payment is never processed twice; promo/loyalty redemption now takes row locks (SELECT FOR UPDATE) so a balance cannot be double-spent under concurrent clicks (S-51).",
    "F-11": "Vite build optimisation at a basic level; instant-load placeholders and skeleton price cells added on the booking and widget pages so the first paint is never empty (S-48).",
    "F-08": "Per-hotel daily quotas; cheap model by default; tool-call limits; response caching; usage tracking now durably stored in Postgres so it survives a Redis outage (PR #63).",
}
for fid, text in prot_updates.items():
    r = find_row(ws, 1, fid)
    if r:
        set_cell(ws, r, PROT, text)

# ---------------------------------------------------------------------------
# SHEET 6 — Team & Hiring : refresh counts (45 -> 58) and an example
# ---------------------------------------------------------------------------
ws = wb["6. Team & Hiring"]
for r in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        v = ws.cell(r, col).value
        if isinstance(v, str) and "45 issues" in v:
            ws.cell(r, col).value = v.replace("45 issues", "58 issues")
        elif isinstance(v, str) and "across 45 issues" in v:
            ws.cell(r, col).value = v.replace("across 45 issues", "across 58 issues")

# ---------------------------------------------------------------------------
# SHEET 8 — Silent Bugs : cite the new real crash fixes
# ---------------------------------------------------------------------------
ws = wb["8. Silent Bugs & Stability"]
for r in range(1, ws.max_row + 1):
    a = ws.cell(r, 1).value
    if isinstance(a, str) and a.startswith("Query returns more rows"):
        ws.cell(r, 2).value = ("An endpoint that assumes one row gets several and throws a 500 only for certain data — "
                               "invisible until that exact data appears. We hit exactly this in the loyalty check (S-46).")
    if isinstance(a, str) and a.startswith("Uncaught frontend error"):
        ws.cell(r, 2).value = ("One failed fetch (or an unsafe bundling change) turns the entire screen white mid-booking — "
                               "the guest just leaves. We hit this twice and fixed both (S-36, S-48).")

# ---------------------------------------------------------------------------
# SHEET 1 — Executive Summary : refresh numbers + 'what's new' block
# ---------------------------------------------------------------------------
ws = wb["1. Executive Summary"]

def replace_in_summary(old_sub, newtext):
    """Find a Topic cell (col A) and overwrite its Detail (col B)."""
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value).strip() == old_sub:
            ws.cell(r, 2).value = newtext
            return r
    return None

replace_in_summary("Sheet 2 — Problems Solved",
    "58 issues fixed so far (6 Critical, 30 High, 22 Medium). Each issue shows what would have happened to the business if it had reached production.")
replace_in_summary("Sheet 3 — Features Built",
    "23 major product features delivered alongside the bug fixing — multi-property chains, loyalty wallet, OTA rate intelligence, AI concierge, real-time updates, encrypted secrets, and the new Revenue Engine (dynamic pricing, abandoned-booking recovery, chain upsell broadcast, seasonal promotions, OTA-style booking widget).")
replace_in_summary("Issues fixed till date",
    "58 (source: work_log_tracker.csv — engineering tasks logged up to 16-Jun-2026; merged across PRs up to #90)")
replace_in_summary("Critical issues fixed",
    "6 — payment secret key leak, payment amount tampering, permanent admin token, AI assistant crash, staff privilege escalation, and the booking-widget code-execution (RCE/XSS) sink.")
replace_in_summary("Automated tests added",
    "Close to 300 automated test cases (the 277 from the June audit, plus new tests for the anti-automation booking token, the recovery scheduler, chain features and public-booking promotions). They run on every code change (CI/CD), with dependency vulnerability (CVE) scanning.")

# Append a 'WHAT'S NEW SINCE THE LAST REPORT' block before the closing message.
# Find the 'ONE-LINE MESSAGE FOR THE CEO' header row to insert before it.
insert_at = None
for r in range(1, ws.max_row + 1):
    if str(ws.cell(r, 1).value).strip() == "ONE-LINE MESSAGE FOR THE CEO":
        insert_at = r
        break
if insert_at:
    ws.insert_rows(insert_at, amount=4)
    a = ws.cell(insert_at, 1, "WHAT'S NEW SINCE 11-JUN")
    a.font = body(bold=True); a.fill = fill(GREEN); a.alignment = TOPWRAP; a.border = BORDER
    ws.cell(insert_at, 2, "").fill = fill(GREEN)
    rows = [
        ("Revenue Engine shipped", "Dynamic pricing (occupancy/lead-time rules wired to real charges), abandoned-booking recovery (WhatsApp + email, hourly scheduler), and a loyalty wallet with stay offers and upsell hub — a whole new revenue surface (PR #80)."),
        ("Chain & promotions", "Chain-wide upsell broadcast, package validity windows, and seasonal auto-apply promotions at hotel and chain level (PR #88)."),
        ("Stability & crash fixes", "Fixed a blank-white-screen crash on deploy (S-48), two booking-flow 500 crashes (S-46, S-47), a dead recovery scheduler (S-49), broken deploys/migrations (S-54, S-56), and removed a widget code-execution security hole (S-50)."),
    ]
    for i, (topic, detail) in enumerate(rows, start=1):
        rr = insert_at + i
        ca = ws.cell(rr, 1, topic); ca.font = body(); ca.alignment = TOPWRAP; ca.border = BORDER
        cb = ws.cell(rr, 2, detail); cb.font = body(); cb.alignment = TOPWRAP; cb.border = BORDER

# Update the bottom-line message
for r in range(1, ws.max_row + 1):
    if str(ws.cell(r, 1).value).strip() == "Bottom line":
        ws.cell(r, 2).value = (
            "The product is stable today and hoteliers' data is safe — and it is moving fast: 58 issues fixed, 23 features shipped, and a full Revenue Engine added since launch. "
            "The hiring need comes from exactly this momentum: as hotels, traffic and data scale up, each layer — database, backend, frontend — needs a dedicated owner so the speed, safety and crash-proofing stay where they are today. "
            "Sheets 7-9 spell out, in plain language, the ongoing engineering that keeps it that way.")
        break

wb.save(PATH)
print("Saved:", PATH)
print("Sheets & sizes:")
for s in wb.worksheets:
    print(f"  {s.title}: rows={s.max_row}")
